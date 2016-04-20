import json
import os.path
import collections
import xlrd
import csv
from openpyxl.reader.excel import load_workbook


def convert(path, filters=None, export_path=None, compress=False):
    if not os.path.isfile(path):
        raise Exception('unable to find file at %s' % path)
    extension = path[path.rfind('.'):]
    if extension == '.xlsx':
        data = extract_data_from_xlsx(path)
    if extension == '.csv':
        data = extract_data_from_csv(path)
    if extension == '.xls':
        data = extract_data_from_xls(path)
    if compress:
        data = compress_data(data)
    if export_path:
        export_to_json(data, export_path)
    return data


def compress_data(data):
    compressed_data = {
        'fields': data[0].keys(),
        'values': []
    }
    for item in data:
        compressed_data['values'].append(item.values())
    return compressed_data


def extract_data_from_csv(file_path):
    data = []
    column_headers = {}
    csv_file = open(file_path)
    csv_file_reader = csv.reader(csv_file)
    for row_idx, row in enumerate(csv_file_reader):
        row_data = collections.OrderedDict()
        for column, cell_val in enumerate(row):
            # populate column headers
            if row_idx == 0:
                column_headers[column] = cell_val
            # populate data
            else:
                row_data[column_headers[column]] = cell_val
        if row_idx != 0:
            data.append(row_data)
    return data


def extract_data_from_xlsx(file_path):
    data = []
    column_headers = {}
    wb = load_workbook(filename=file_path)
    ws = wb.active
    for row in range(1, ws.max_row + 1):
        row_data = collections.OrderedDict()
        for column in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=row, column=column).value
            # populate column headers
            if row == 1:
                column_headers[column] = cell_val
            # populate data
            else:
                row_data[column_headers[column]] = cell_val
        if row != 1:
            data.append(row_data)
    return data


def extract_data_from_xls(filename):
    book = xlrd.open_workbook(filename)
    sheet = book.sheet_by_index(0)
    data = []
    column_headers = {}
    for row in range(0, sheet.nrows):
        row_data = collections.OrderedDict()
        for column in range(0, sheet.ncols):
            cell_val = sheet.cell_value(row, column)
            # populate column headers
            if row == 0:
                column_headers[column] = cell_val
            # populate data
            else:
                row_data[column_headers[column]] = cell_val
        if row != 0:
            data.append(row_data)
    return data


def export_to_json(data, filename='data.json'):
    if data is not None:
        with open(filename, 'w') as outfile:
            json.dump(data, outfile)
